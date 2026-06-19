from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, Response
from flask_login import login_required, current_user
from database import get_db
from datetime import datetime, timedelta
import io

inventar_bp = Blueprint('inventar', __name__, url_prefix='/inventar')

def get_stoc_locatie(db, locatie_id, produs_id=None):
    """Calculeaza stocul curent la o locatie pentru un produs sau toate produsele."""
    q = """
        SELECT p.id, p.denumire, p.unitate_masura,
        COALESCE((SELECT SUM(dd.cantitate) FROM distributii_detalii dd
                  JOIN distributii d ON dd.distributie_id=d.id
                  WHERE dd.produs_id=p.id AND d.locatie_id=?), 0)
        +
        COALESCE((SELECT SUM(td.cantitate) FROM transferuri_detalii td
                  JOIN transferuri t ON td.transfer_id=t.id
                  WHERE td.produs_id=p.id AND t.locatie_destinatie_id=?), 0)
        +
        COALESCE((SELECT SUM(rd.cantitate) FROM retururi_detalii rd
                  JOIN retururi r ON rd.retur_id=r.id
                  WHERE rd.produs_id=p.id AND r.locatie_id=?), 0) * -1
        -
        COALESCE((SELECT SUM(td2.cantitate) FROM transferuri_detalii td2
                  JOIN transferuri t2 ON td2.transfer_id=t2.id
                  WHERE td2.produs_id=p.id AND t2.locatie_sursa_id=?), 0)
        -
        COALESCE((SELECT SUM(vd.cantitate) FROM vanzari_detalii vd
                  JOIN vanzari_import vi ON vd.import_id=vi.id
                  WHERE vd.produs_id=p.id AND vi.locatie_id=?), 0)
        -
        COALESCE((SELECT SUM(pi.cantitate) FROM pierderi pi
                  WHERE pi.produs_id=p.id AND pi.locatie_id=?), 0)
        as stoc_curent
        FROM produse p WHERE p.activ=1
    """
    params = [locatie_id, locatie_id, locatie_id, locatie_id, locatie_id, locatie_id]
    if produs_id:
        q += " AND p.id=?"
        params.append(produs_id)
    q += " ORDER BY p.denumire"
    return db.execute(q, params).fetchall()


def _calculeaza_date_inventar(db, locatie_id, data_inv, exclude_inv_id=None):
    """
    Returneaza (stoc_initial_map, intrari_map, vanzari_map) pentru o locatie si o data.

    - stoc_initial_map : {produs_id: scriptic} din ultimul inventar finalizat anterior
    - intrari_map      : {produs_id: total}   din distributii in saptamana inventarului
                         (gol daca nu exista distributii)
    - vanzari_map      : {produs_id: total}   din vanzari_import in saptamana inventarului
    """
    # ── Stoc initial: scriptul inventarului anterior finalizat ────────────────
    q_prev = """SELECT id FROM inventar
                WHERE locatie_id=? AND finalizat=1 AND data < ?"""
    params_prev = [locatie_id, data_inv]
    if exclude_inv_id:
        q_prev += " AND id != ?"
        params_prev.append(exclude_inv_id)
    q_prev += " ORDER BY data DESC, id DESC LIMIT 1"
    prev_inv = db.execute(q_prev, params_prev).fetchone()

    stoc_initial_map = {}
    if prev_inv:
        rows = db.execute("""
            SELECT produs_id,
                   (COALESCE(stoc_initial, 0) + COALESCE(intrari, 0)) as scriptic
            FROM inventar_detalii WHERE inventar_id=?
        """, (prev_inv['id'],)).fetchall()
        stoc_initial_map = {r['produs_id']: round(r['scriptic'], 3) for r in rows}

    # ── Intrari: distributii catre locatie pana la data inventarului ──────────
    dist_rows = db.execute("""
        SELECT dd.produs_id, SUM(dd.cantitate) as total
        FROM distributii_detalii dd
        JOIN distributii d ON dd.distributie_id = d.id
        WHERE d.locatie_id = ? AND d.data <= ?
        GROUP BY dd.produs_id
    """, (locatie_id, data_inv)).fetchall()
    intrari_map = {r['produs_id']: round(r['total'], 3) for r in dist_rows}

    # ── Vanzari: cel mai recent raport Carrefour disponibil pentru locatie ────
    # (nu saptamana fixa — Carrefour trimite raportul cu decalaj)
    latest = db.execute("""
        SELECT MAX(data_raportare) as max_data FROM vanzari_import
        WHERE locatie_id = ? AND data_raportare <= ?
    """, (locatie_id, data_inv)).fetchone()

    vanzari_map = {}
    if latest and latest['max_data']:
        vanz_rows = db.execute("""
            SELECT vd.produs_id, SUM(vd.cantitate) as total
            FROM vanzari_detalii vd
            JOIN vanzari_import vi ON vd.import_id = vi.id
            WHERE vi.locatie_id = ? AND vi.data_raportare = ?
              AND vd.produs_id IS NOT NULL
            GROUP BY vd.produs_id
        """, (locatie_id, latest['max_data'])).fetchall()
        vanzari_map = {r['produs_id']: round(r['total'], 3) for r in vanz_rows}

    return stoc_initial_map, intrari_map, vanzari_map


@inventar_bp.route('/')
@login_required
def index():
    db = get_db()
    q = """SELECT i.*, l.nume as locatie_nume,
           COUNT(id2.id) as nr_produse,
           u.nume_complet as utilizator_nume
           FROM inventar i LEFT JOIN locatii l ON i.locatie_id=l.id
           LEFT JOIN inventar_detalii id2 ON id2.inventar_id=i.id
           LEFT JOIN utilizatori u ON i.utilizator_id=u.id"""
    params = []
    if not current_user.is_manager() and current_user.locatie_id:
        q += " WHERE i.locatie_id=?"
        params.append(current_user.locatie_id)
    q += " GROUP BY i.id ORDER BY i.data DESC, i.id DESC LIMIT 50"
    inventare = db.execute(q, params).fetchall()
    db.close()
    return render_template('inventar/index.html', inventare=inventare)


@inventar_bp.route('/nou', methods=['GET', 'POST'])
@login_required
def nou():
    db = get_db()
    if request.method == 'POST':
        data = request.form.get('data')
        locatie_id = request.form.get('locatie_id')
        if not current_user.is_manager():
            locatie_id = current_user.locatie_id
        observatii = request.form.get('observatii', '').strip()
        if not data or not locatie_id:
            flash('Data si locatia sunt obligatorii.', 'danger')
        else:
            existing = db.execute("""SELECT id FROM inventar WHERE locatie_id=? AND finalizat=0 AND data=?""",
                                  (locatie_id, data)).fetchone()
            if existing:
                flash('Exista deja un inventar nefinalizat pentru aceasta locatie si data.', 'warning')
                return redirect(url_for('inventar.completeaza', id=existing['id']))

            cur = db.execute("""INSERT INTO inventar (data, locatie_id, observatii, utilizator_id)
                VALUES (?,?,?,?)""", (data, locatie_id, observatii, current_user.id))
            inv_id = cur.lastrowid

            # Populeaza din baza de date:
            # - Stoc Initial = 0 (se introduce manual, cantarire fizica lunea)
            # - Intrari = din distributii pentru saptamana inventarului
            # - Vanzari Reale = din rapoarte Carrefour pentru saptamana inventarului
            _, intr_map, vanz_map = _calculeaza_date_inventar(db, locatie_id, data)
            all_prod_ids = set(intr_map) | set(vanz_map)

            for prod_id in all_prod_ids:
                intr = intr_map.get(prod_id, 0) or 0
                vanz = vanz_map.get(prod_id, 0) or 0
                db.execute("""INSERT INTO inventar_detalii
                    (inventar_id, produs_id, cantitate_sistem, cantitate_fizica, stoc_initial, intrari)
                    VALUES (?,?,?,?,?,?)""",
                    (inv_id, prod_id, intr, vanz, 0, intr))

            # Adauga produsele care nu apar in vanzari/intrari (pentru completare manuala)
            produse_toate = db.execute("SELECT id FROM produse WHERE activ=1").fetchall()
            for p in produse_toate:
                if p['id'] not in all_prod_ids:
                    db.execute("""INSERT INTO inventar_detalii
                        (inventar_id, produs_id, cantitate_sistem, cantitate_fizica, stoc_initial, intrari)
                        VALUES (?,?,?,?,?,?)""",
                        (inv_id, p['id'], 0, 0, 0, 0))

            db.commit()
            flash('Inventar creat si populat din baza de date.', 'success')
            return redirect(url_for('inventar.completeaza', id=inv_id))

    if current_user.is_manager():
        locatii = db.execute("SELECT * FROM locatii WHERE activa=1 ORDER BY nume").fetchall()
    else:
        locatii = db.execute("SELECT * FROM locatii WHERE id=?", (current_user.locatie_id,)).fetchall()
    db.close()
    return render_template('inventar/nou.html', locatii=locatii)


@inventar_bp.route('/api/date-sistem/<int:inv_id>')
@login_required
def api_date_sistem(inv_id):
    """Returneaza datele din sistem (stoc_initial, intrari, vanzari) per produs pentru un inventar."""
    db = get_db()
    inv = db.execute("SELECT * FROM inventar WHERE id=?", (inv_id,)).fetchone()
    if not inv:
        db.close()
        return jsonify({'error': 'Inventar negasit'}), 404

    _, intr_map, vanz_map = _calculeaza_date_inventar(
        db, inv['locatie_id'], inv['data'], exclude_inv_id=inv_id
    )

    # Stoc Initial nu se preia din DB - il introduce utilizatorul manual (cantarire lunea)
    result = {}
    all_ids = set(intr_map) | set(vanz_map)
    for pid in all_ids:
        result[str(pid)] = {
            'intrari': intr_map.get(pid, 0) or 0,
            'vanzari': vanz_map.get(pid, 0) or 0,
        }

    db.close()
    return jsonify(result)


@inventar_bp.route('/completeaza/<int:id>', methods=['GET', 'POST'])
@login_required
def completeaza(id):
    db = get_db()
    inv = db.execute("""SELECT i.*, l.nume as locatie_nume FROM inventar i
        LEFT JOIN locatii l ON i.locatie_id=l.id WHERE i.id=?""", (id,)).fetchone()
    if not inv:
        flash('Inventar negasit.', 'danger')
        return redirect(url_for('inventar.index'))

    if request.method == 'POST':
        actiune = request.form.get('actiune', 'salveaza')
        detalii_ids = request.form.getlist('detaliu_id')
        stoc_initiale = request.form.getlist('stoc_initial')
        intrari_list = request.form.getlist('intrari')
        cantitati = request.form.getlist('cantitate_fizica')
        for det_id, si, intr, cant in zip(detalii_ids, stoc_initiale, intrari_list, cantitati):
            try:
                db.execute("""UPDATE inventar_detalii
                    SET stoc_initial=?, intrari=?, cantitate_fizica=?
                    WHERE id=?""",
                    (float(si or 0), float(intr or 0), float(cant or 0), int(det_id)))
            except:
                pass
        if actiune == 'finalizeaza':
            db.execute("UPDATE inventar SET finalizat=1 WHERE id=?", (id,))
            flash('Inventar finalizat.', 'success')
        else:
            flash('Inventar salvat.', 'success')
        db.commit()
        if actiune == 'finalizeaza':
            return redirect(url_for('inventar.index'))
        return redirect(url_for('inventar.completeaza', id=id))

    detalii_raw = db.execute("""
        SELECT id2.*, p.denumire, p.unitate_masura
        FROM inventar_detalii id2
        JOIN produse p ON id2.produs_id=p.id
        WHERE id2.inventar_id=?
        ORDER BY p.denumire
    """, (id,)).fetchall()
    detalii = []
    for d in detalii_raw:
        d = dict(d)
        d['scriptic'] = round((d.get('stoc_initial') or 0) + (d.get('intrari') or 0), 3)
        d['diferenta'] = round(d['scriptic'] - (d.get('cantitate_fizica') or 0), 3)
        detalii.append(d)
    db.close()
    return render_template('inventar/completeaza.html', inv=inv, detalii=detalii)


@inventar_bp.route('/sterge/<int:id>', methods=['POST'])
@login_required
def sterge(id):
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('inventar.index'))
    db = get_db()
    db.execute("DELETE FROM inventar WHERE id=?", (id,))
    db.commit()
    db.close()
    flash('Inventar sters.', 'success')
    return redirect(url_for('inventar.index'))


@inventar_bp.route('/detalii/<int:id>')
@login_required
def detalii(id):
    db = get_db()
    inv = db.execute("""SELECT i.*, l.nume as locatie_nume FROM inventar i
        LEFT JOIN locatii l ON i.locatie_id=l.id WHERE i.id=?""", (id,)).fetchone()
    detalii_raw = db.execute("""
        SELECT id2.*, p.denumire, p.unitate_masura
        FROM inventar_detalii id2
        JOIN produse p ON id2.produs_id=p.id
        WHERE id2.inventar_id=?
        ORDER BY p.denumire
    """, (id,)).fetchall()
    detalii = []
    for d in detalii_raw:
        d = dict(d)
        d['scriptic'] = round((d.get('stoc_initial') or 0) + (d.get('intrari') or 0), 3)
        d['diferenta'] = round(d['scriptic'] - (d.get('cantitate_fizica') or 0), 3)
        detalii.append(d)
    db.close()
    return render_template('inventar/detalii.html', inv=inv, detalii=detalii)


# ─── INVENTAR AUTOMAT (bazat pe stoc_initial_locatii) ─────────────────────────

def _calcul_raport(db, locatie_id, data_si, data_sf):
    """
    Calculeaza raportul inventar automat pentru o locatie si doua date de stoc.
    data_si = stoc initial (saptamana trecuta), data_sf = stoc fizic curent.
    """
    # Stoc fizic la data_sf
    sf_rows = db.execute("""
        SELECT p.id as produs_id, p.denumire, p.unitate_masura as um,
               si.cantitate as stoc_fizic
        FROM stoc_initial_locatii si
        JOIN produse p ON p.id = si.produs_id
        WHERE si.locatie_id = ? AND si.data = ?
    """, (locatie_id, data_sf)).fetchall()

    # Stoc initial la data_si
    si_map = {}
    if data_si:
        si_rows = db.execute("""
            SELECT produs_id, cantitate FROM stoc_initial_locatii
            WHERE locatie_id = ? AND data = ?
        """, (locatie_id, data_si)).fetchall()
        si_map = {r['produs_id']: r['cantitate'] for r in si_rows}

    date_cond = ""
    date_params = [locatie_id]
    if data_si and data_sf:
        date_cond = "AND d.data > ? AND d.data <= ?"
        date_params += [data_si, data_sf]
    elif data_sf:
        date_cond = "AND d.data <= ?"
        date_params += [data_sf]

    # Intrari = distributii de la depozit catre locatie
    intr_rows = db.execute(f"""
        SELECT dd.produs_id, SUM(dd.cantitate) as total
        FROM distributii_detalii dd
        JOIN distributii d ON dd.distributie_id = d.id
        WHERE d.locatie_id = ? {date_cond}
        GROUP BY dd.produs_id
    """, date_params).fetchall()
    intr_map = {r['produs_id']: r['total'] for r in intr_rows}

    # Vanzari Carrefour
    vc_params = [locatie_id]
    vc_cond = ""
    if data_si and data_sf:
        vc_cond = "AND vi.data_raportare > ? AND vi.data_raportare <= ?"
        vc_params += [data_si, data_sf]
    elif data_sf:
        vc_cond = "AND vi.data_raportare <= ?"
        vc_params += [data_sf]
    vanz_rows = db.execute(f"""
        SELECT vd.produs_id, SUM(vd.cantitate) as total
        FROM vanzari_detalii vd
        JOIN vanzari_import vi ON vd.import_id = vi.id
        WHERE vi.locatie_id = ? {vc_cond}
        GROUP BY vd.produs_id
    """, vc_params).fetchall()
    vanz_map = {r['produs_id']: r['total'] for r in vanz_rows}

    # Transferuri catre locatie
    tr_params_in = [locatie_id]
    tr_params_out = [locatie_id]
    tr_cond = ""
    if data_si and data_sf:
        tr_cond = "AND t.data > ? AND t.data <= ?"
        tr_params_in += [data_si, data_sf]
        tr_params_out += [data_si, data_sf]
    elif data_sf:
        tr_cond = "AND t.data <= ?"
        tr_params_in += [data_sf]
        tr_params_out += [data_sf]
    tin_rows = db.execute(f"""
        SELECT td.produs_id, SUM(td.cantitate) as total
        FROM transferuri_detalii td
        JOIN transferuri t ON td.transfer_id = t.id
        WHERE t.locatie_destinatie_id = ? {tr_cond}
        GROUP BY td.produs_id
    """, tr_params_in).fetchall() if tr_params_in else []
    tin_map = {r['produs_id']: r['total'] for r in tin_rows}

    tout_rows = db.execute(f"""
        SELECT td.produs_id, SUM(td.cantitate) as total
        FROM transferuri_detalii td
        JOIN transferuri t ON td.transfer_id = t.id
        WHERE t.locatie_sursa_id = ? {tr_cond}
        GROUP BY td.produs_id
    """, tr_params_out).fetchall() if tr_params_out else []
    tout_map = {r['produs_id']: r['total'] for r in tout_rows}

    # Retururi
    ret_params = [locatie_id]
    ret_cond = ""
    if data_si and data_sf:
        ret_cond = "AND r.data > ? AND r.data <= ?"
        ret_params += [data_si, data_sf]
    elif data_sf:
        ret_cond = "AND r.data <= ?"
        ret_params += [data_sf]
    ret_rows = db.execute(f"""
        SELECT rd.produs_id, SUM(rd.cantitate) as total
        FROM retururi_detalii rd
        JOIN retururi r ON rd.retur_id = r.id
        WHERE r.locatie_id = ? {ret_cond}
        GROUP BY rd.produs_id
    """, ret_params).fetchall()
    ret_map = {r['produs_id']: r['total'] for r in ret_rows}

    detalii = []
    total_stoc_initial = total_intrari = total_vanzari = 0
    total_scriptic = total_stoc_fizic = total_plus = total_minus = 0

    for r in sf_rows:
        pid = r['produs_id']
        si_val  = round(si_map.get(pid, 0) or 0, 3)
        intr    = round(intr_map.get(pid, 0) or 0, 3)
        vanz    = round(vanz_map.get(pid, 0) or 0, 3)
        tin     = round(tin_map.get(pid, 0) or 0, 3)
        tout    = round(tout_map.get(pid, 0) or 0, 3)
        ret     = round(ret_map.get(pid, 0) or 0, 3)
        sf_val  = round(r['stoc_fizic'] or 0, 3)

        scriptic  = round(si_val + intr + tin - tout - ret - vanz, 3)
        diferenta = round(sf_val - scriptic, 3)

        detalii.append({
            'produs_id': pid, 'denumire': r['denumire'], 'um': r['um'],
            'stoc_initial': si_val, 'intrari': intr, 'vanzari': vanz,
            'transferuri_in': tin, 'transferuri_out': tout, 'retururi': ret,
            'scriptic': scriptic, 'stoc_fizic': sf_val, 'diferenta': diferenta,
        })

        total_stoc_initial += si_val
        total_intrari      += intr
        total_vanzari      += vanz
        total_scriptic     += scriptic
        total_stoc_fizic   += sf_val
        if diferenta > 0:
            total_plus += diferenta
        else:
            total_minus += diferenta

    detalii.sort(key=lambda x: x['denumire'])
    totale = {
        'stoc_initial': round(total_stoc_initial, 3),
        'intrari':      round(total_intrari, 3),
        'vanzari':      round(total_vanzari, 3),
        'scriptic':     round(total_scriptic, 3),
        'stoc_fizic':   round(total_stoc_fizic, 3),
        'diferenta':    round(total_stoc_fizic - total_scriptic, 3),
        'plus':         round(total_plus, 3),
        'minus':        round(total_minus, 3),
    }
    return detalii, totale


@inventar_bp.route('/raport')
@login_required
def raport_index():
    db = get_db()
    rows = db.execute("""
        SELECT l.id, l.nume,
               MAX(si.data) as ultima_data_stoc,
               COUNT(DISTINCT si.data) as nr_importuri,
               (SELECT COUNT(*) FROM vanzari_import vi WHERE vi.locatie_id=l.id) as nr_vanzari,
               (SELECT COUNT(*) FROM distributii d WHERE d.locatie_id=l.id) as nr_intrari
        FROM locatii l
        LEFT JOIN stoc_initial_locatii si ON si.locatie_id = l.id
        WHERE l.activa = 1
        GROUP BY l.id
        ORDER BY l.nume
    """).fetchall()
    db.close()
    return render_template('inventar/raport_index.html', locatii=[dict(r) for r in rows])


@inventar_bp.route('/raport/<int:locatie_id>')
@login_required
def raport(locatie_id):
    db = get_db()
    locatie = db.execute("SELECT * FROM locatii WHERE id=?", (locatie_id,)).fetchone()
    if not locatie:
        flash('Locatie negasita.', 'danger')
        db.close()
        return redirect(url_for('inventar.raport_index'))

    dates_stoc = [r['data'] for r in db.execute("""
        SELECT DISTINCT data FROM stoc_initial_locatii
        WHERE locatie_id=? ORDER BY data DESC
    """, (locatie_id,)).fetchall()]

    data_si = request.args.get('data_si', '').strip()
    data_sf = request.args.get('data_sf', dates_stoc[0] if dates_stoc else '').strip()

    detalii, totale = [], {}
    if data_sf:
        detalii, totale = _calcul_raport(db, locatie_id, data_si or None, data_sf)

    db.close()
    return render_template('inventar/raport.html',
        locatie=locatie, dates_stoc=dates_stoc,
        data_si=data_si, data_sf=data_sf,
        detalii=detalii, totale=totale)


@inventar_bp.route('/raport/<int:locatie_id>/export')
@login_required
def raport_export(locatie_id):
    db = get_db()
    locatie = db.execute("SELECT * FROM locatii WHERE id=?", (locatie_id,)).fetchone()
    data_si = request.args.get('data_si', '').strip()
    data_sf = request.args.get('data_sf', '').strip()

    if not locatie or not data_sf:
        flash('Date insuficiente pentru export.', 'danger')
        db.close()
        return redirect(url_for('inventar.raport_index'))

    detalii, totale = _calcul_raport(db, locatie_id, data_si or None, data_sf)
    db.close()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Inventar {locatie['nume']}"
        headers = ['Produs', 'UM', 'Stoc Initial', 'Intrari', 'Vanzari', 'Scriptic', 'Stoc Fizic', 'Diferenta']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for d in detalii:
            ws.append([d['denumire'], d['um'], d['stoc_initial'], d['intrari'],
                       d['vanzari'], d['scriptic'], d['stoc_fizic'], d['diferenta']])
        ws.append(['TOTAL', '', totale.get('stoc_initial', 0), totale.get('intrari', 0),
                   totale.get('vanzari', 0), totale.get('scriptic', 0),
                   totale.get('stoc_fizic', 0), totale.get('diferenta', 0)])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"inventar_{locatie['nume']}_{data_sf}.xlsx".replace(' ', '_')
        return Response(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'})
    except ImportError:
        flash('openpyxl nu este instalat.', 'danger')
        return redirect(url_for('inventar.raport', locatie_id=locatie_id))
