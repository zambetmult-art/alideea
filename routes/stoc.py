from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from database import get_db
from datetime import datetime
import openpyxl
import io
import re
import difflib

stoc_bp = Blueprint('stoc', __name__, url_prefix='/stoc')


@stoc_bp.route('/')
@login_required
def index():
    db = get_db()
    locatii = db.execute("SELECT * FROM locatii WHERE activa=1 ORDER BY nume").fetchall()

    loc_filter  = request.args.get('locatie_id', '').strip()
    data_filter = request.args.get('data', '').strip()

    # Submisii inventar in asteptare (trimise de angajati via link)
    pending = db.execute("""
        SELECT s.*, l.nume as locatie_nume,
               COUNT(sd.id) as nr_produse
        FROM submisii s
        LEFT JOIN locatii l ON s.locatie_id = l.id
        LEFT JOIN submisii_detalii sd ON sd.submisie_id = s.id
        WHERE s.tip = 'inventar' AND s.status = 'pending'
        GROUP BY s.id
        ORDER BY s.data_submisie DESC
    """).fetchall()

    # Stocuri importate/confirmate, grupate pe (locatie, data)
    q_where  = "1=1"
    q_params = []
    if loc_filter:
        try:
            q_params.append(int(loc_filter))
            q_where += " AND si.locatie_id = ?"
        except ValueError:
            pass
    if data_filter:
        q_where += " AND si.data = ?"
        q_params.append(data_filter)

    snapshots = db.execute(f"""
        SELECT si.locatie_id, si.data,
               l.nume as locatie_nume,
               COUNT(si.id) as nr_produse,
               MAX(si.created_at) as creat_la,
               u.nume_complet as creat_de_nume
        FROM stoc_initial_locatii si
        JOIN locatii l ON si.locatie_id = l.id
        LEFT JOIN utilizatori u ON si.utilizator_id = u.id
        WHERE {q_where}
        GROUP BY si.locatie_id, si.data
        ORDER BY si.data DESC, l.nume
    """, q_params).fetchall()

    db.close()
    return render_template('stoc/index.html',
        locatii=locatii,
        pending=[dict(r) for r in pending],
        snapshots=[dict(r) for r in snapshots],
        loc_filter=loc_filter,
        data_filter=data_filter,
        today=datetime.now().strftime('%Y-%m-%d'),
    )


@stoc_bp.route('/snapshot/<int:locatie_id>/<data>')
@login_required
def snapshot_detalii(locatie_id, data):
    db = get_db()
    locatie = db.execute("SELECT * FROM locatii WHERE id=?", (locatie_id,)).fetchone()
    if not locatie:
        flash('Locatie negasita.', 'danger')
        db.close()
        return redirect(url_for('stoc.index'))
    produse = db.execute("""
        SELECT p.cod_articol, p.denumire, p.unitate_masura,
               c.nume as categorie, si.cantitate, si.created_at
        FROM stoc_initial_locatii si
        JOIN produse p ON p.id = si.produs_id
        LEFT JOIN categorii c ON c.id = p.categorie_id
        WHERE si.locatie_id = ? AND si.data = ?
        ORDER BY c.nume, p.denumire
    """, (locatie_id, data)).fetchall()
    db.close()
    return render_template('stoc/snapshot.html',
        locatie=locatie, data=data,
        produse=[dict(r) for r in produse]
    )


@stoc_bp.route('/snapshot/<int:locatie_id>/<data>/sterge', methods=['POST'])
@login_required
def sterge_snapshot(locatie_id, data):
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('stoc.index'))
    db = get_db()
    db.execute("DELETE FROM stoc_initial_locatii WHERE locatie_id=? AND data=?", (locatie_id, data))
    db.commit()
    db.close()
    flash('Stoc sters.', 'success')
    return redirect(url_for('stoc.index'))


@stoc_bp.route('/calculat')
@login_required
def calculat():
    db = get_db()
    locatii = db.execute("SELECT * FROM locatii WHERE activa=1 ORDER BY nume").fetchall()
    locatie_id = request.args.get('locatie_id', 'central')

    if locatie_id == 'central':
        stoc = db.execute("""
            SELECT p.id, p.cod_articol, p.denumire, p.unitate_masura, p.stoc_minim,
                   c.nume as categorie_nume,
                   COALESCE(SUM(CASE WHEN src='intrare' THEN cant ELSE -cant END), 0) as stoc_curent
            FROM produse p
            LEFT JOIN categorii c ON p.categorie_id=c.id
            LEFT JOIN (
                SELECT produs_id, cantitate as cant, 'intrare' as src FROM intrari_detalii
                UNION ALL
                SELECT produs_id, cantitate, 'iesire' FROM distributii_detalii
            ) mv ON mv.produs_id = p.id
            WHERE p.activ=1
            GROUP BY p.id
            ORDER BY c.nume, p.denumire
        """).fetchall()
        titlu = "Stoc Central"
    else:
        try:
            lid = int(locatie_id)
        except:
            lid = 0
        locatie = db.execute("SELECT * FROM locatii WHERE id=?", (lid,)).fetchone()
        titlu = f"Stoc - {locatie['nume']}" if locatie else "Stoc Locatie"

        stoc = db.execute("""
            SELECT p.id, p.cod_articol, p.denumire, p.unitate_masura, p.stoc_minim,
                   c.nume as categorie_nume,
                   COALESCE(si.cantitate, 0) as stoc_initial,
                   si.data as data_stoc_initial,
                   COALESCE(si.cantitate, 0)
                   + COALESCE((SELECT SUM(dd.cantitate) FROM distributii_detalii dd
                              JOIN distributii d ON dd.distributie_id=d.id
                              WHERE dd.produs_id=p.id AND d.locatie_id=?
                                AND d.data >= COALESCE(si.data, '0000-00-00')), 0)
                   + COALESCE((SELECT SUM(td.cantitate) FROM transferuri_detalii td
                                JOIN transferuri t ON td.transfer_id=t.id
                                WHERE td.produs_id=p.id AND t.locatie_destinatie_id=?
                                  AND t.data >= COALESCE(si.data, '0000-00-00')), 0)
                   - COALESCE((SELECT SUM(rd.cantitate) FROM retururi_detalii rd
                                JOIN retururi r ON rd.retur_id=r.id
                                WHERE rd.produs_id=p.id AND r.locatie_id=?
                                  AND r.data >= COALESCE(si.data, '0000-00-00')), 0)
                   - COALESCE((SELECT SUM(td2.cantitate) FROM transferuri_detalii td2
                                JOIN transferuri t2 ON td2.transfer_id=t2.id
                                WHERE td2.produs_id=p.id AND t2.locatie_sursa_id=?
                                  AND t2.data >= COALESCE(si.data, '0000-00-00')), 0)
                   - COALESCE((SELECT SUM(vd.cantitate) FROM vanzari_detalii vd
                                JOIN vanzari_import vi ON vd.import_id=vi.id
                                WHERE vd.produs_id=p.id AND vi.locatie_id=?
                                  AND vi.data_raportare >= COALESCE(si.data, '0000-00-00')), 0)
                   - COALESCE((SELECT SUM(pi.cantitate) FROM pierderi pi
                                WHERE pi.produs_id=p.id AND pi.locatie_id=?
                                  AND pi.data >= COALESCE(si.data, '0000-00-00')), 0)
                   as stoc_curent
            FROM produse p
            LEFT JOIN categorii c ON p.categorie_id=c.id
            LEFT JOIN stoc_initial_locatii si ON si.produs_id=p.id AND si.locatie_id=?
            WHERE p.activ=1
            ORDER BY c.nume, p.denumire
        """, (lid, lid, lid, lid, lid, lid, lid)).fetchall()

    db.close()
    return render_template('stoc/calculat.html',
        stoc=stoc, locatii=locatii,
        locatie_id=locatie_id, titlu=titlu)


def _parse_excel_stoc(file_storage):
    continut = file_storage.read()
    wb = openpyxl.load_workbook(io.BytesIO(continut), read_only=True, data_only=True)
    ws = wb.active
    rezultat = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        ident = str(row[0]).strip()
        if not ident:
            continue
        raw_qty = row[1] if len(row) > 1 else None
        if raw_qty is None:
            continue
        try:
            qty = float(str(raw_qty).replace(',', '.').strip())
        except (ValueError, TypeError):
            continue
        rezultat.append((ident, qty))
    wb.close()
    return rezultat


def _match_produse(db, rows_excel):
    produse = db.execute("SELECT id, denumire, cod_articol, cod_ean FROM produse WHERE activ=1").fetchall()
    by_name = {p['denumire'].lower().strip(): p for p in produse}
    by_cod  = {p['cod_articol'].strip(): p for p in produse if p['cod_articol']}
    by_ean  = {str(p['cod_ean']).strip(): p for p in produse if p['cod_ean']}

    matched   = []
    unmatched = []

    for ident, qty in rows_excel:
        p = by_ean.get(ident) or by_cod.get(ident) or by_name.get(ident.lower())
        if not p:
            ident_low = ident.lower()
            for key, pp in by_name.items():
                if ident_low in key or key in ident_low:
                    p = pp
                    break
        if p:
            matched.append({'produs_id': p['id'], 'produs': p['denumire'], 'cantitate': qty})
        else:
            unmatched.append(ident)

    return matched, unmatched


def _parse_excel_stoc_toate(file_storage):
    continut = file_storage.read()
    wb = openpyxl.load_workbook(io.BytesIO(continut), read_only=True, data_only=True)
    ws = wb.active
    rezultat = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None or len(row) < 3:
            continue
        locatie_str = str(row[0]).strip()
        produs_str = str(row[1]).strip() if row[1] is not None else ''
        raw_qty = row[2]
        if not locatie_str or not produs_str or raw_qty is None:
            continue
        try:
            qty = float(str(raw_qty).replace(',', '.').strip())
        except (ValueError, TypeError):
            continue
        rezultat.append((locatie_str, produs_str, qty))
    wb.close()
    return rezultat


def _match_produse_locatii(db, rows_excel):
    locatii = db.execute("SELECT id, nume FROM locatii WHERE activa=1").fetchall()
    produse = db.execute("SELECT id, denumire, cod_articol, cod_ean FROM produse WHERE activ=1").fetchall()

    loc_by_nume = {l['nume'].lower().strip(): l for l in locatii}
    by_name = {p['denumire'].lower().strip(): p for p in produse}
    by_cod  = {p['cod_articol'].strip(): p for p in produse if p['cod_articol']}
    by_ean  = {str(p['cod_ean']).strip(): p for p in produse if p['cod_ean']}

    matched   = []
    unmatched = []

    for locatie_str, produs_str, qty in rows_excel:
        loc = loc_by_nume.get(locatie_str.lower())
        if not loc:
            loc_low = locatie_str.lower()
            for key, l in loc_by_nume.items():
                if loc_low in key or key in loc_low:
                    loc = l
                    break

        p = by_ean.get(produs_str) or by_cod.get(produs_str) or by_name.get(produs_str.lower())
        if not p:
            ident_low = produs_str.lower()
            for key, pp in by_name.items():
                if ident_low in key or key in ident_low:
                    p = pp
                    break

        if loc and p:
            matched.append({
                'locatie_id': loc['id'], 'locatie': loc['nume'],
                'produs_id': p['id'], 'produs': p['denumire'],
                'cantitate': qty,
            })
        elif not loc:
            unmatched.append(f'{locatie_str} / {produs_str} — locatie negasita')
        else:
            unmatched.append(f'{locatie_str} / {produs_str} — produs negasit')

    return matched, unmatched


@stoc_bp.route('/import-toate', methods=['GET', 'POST'])
@login_required
def import_toate():
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('stoc.index'))

    db = get_db()
    matched   = []
    unmatched = []
    importat  = False

    if request.method == 'POST':
        fisier = request.files.get('fisier')
        data_stoc = request.form.get('data_stoc', '').strip() or datetime.now().strftime('%Y-%m-%d')
        if not fisier or not fisier.filename:
            flash('Selecteaza un fisier Excel.', 'danger')
        else:
            try:
                rows_excel = _parse_excel_stoc_toate(fisier)
                matched, unmatched = _match_produse_locatii(db, rows_excel)

                if matched:
                    for m in matched:
                        db.execute("""
                            INSERT OR REPLACE INTO stoc_initial_locatii
                                (locatie_id, produs_id, cantitate, data, utilizator_id)
                            VALUES (?,?,?,?,?)
                        """, (m['locatie_id'], m['produs_id'], m['cantitate'], data_stoc, current_user.id))
                    db.commit()
                    importat = True
                    flash(f'{len(matched)} randuri importate cu succes pentru data {data_stoc}.', 'success')
                else:
                    flash('Niciun rand nu a putut fi potrivit. Verifica formatul fisierului.', 'warning')

            except Exception as e:
                flash(f'Eroare la citirea fisierului: {e}', 'danger')

    db.close()
    return render_template('stoc/import_toate.html', matched=matched, unmatched=unmatched, importat=importat)


@stoc_bp.route('/import/<int:locatie_id>', methods=['GET', 'POST'])
@login_required
def import_excel(locatie_id):
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('stoc.index'))

    db = get_db()
    locatie = db.execute("SELECT * FROM locatii WHERE id=?", (locatie_id,)).fetchone()
    if not locatie:
        db.close()
        flash('Locatie negasita.', 'danger')
        return redirect(url_for('stoc.index'))

    matched   = []
    unmatched = []
    importat  = False

    if request.method == 'POST':
        fisier = request.files.get('fisier')
        data_stoc = request.form.get('data_stoc', '').strip() or datetime.now().strftime('%Y-%m-%d')
        if not fisier or not fisier.filename:
            flash('Selecteaza un fisier Excel.', 'danger')
        else:
            try:
                rows_excel = _parse_excel_stoc(fisier)
                matched, unmatched = _match_produse(db, rows_excel)

                if matched:
                    for m in matched:
                        db.execute("""
                            INSERT OR REPLACE INTO stoc_initial_locatii
                                (locatie_id, produs_id, cantitate, data, utilizator_id)
                            VALUES (?,?,?,?,?)
                        """, (locatie_id, m['produs_id'], m['cantitate'], data_stoc, current_user.id))
                    db.commit()
                    importat = True
                    flash(f'{len(matched)} produse importate pentru {locatie["nume"]} / {data_stoc}.', 'success')
                else:
                    flash('Niciun produs nu a putut fi potrivit. Verifica formatul fisierului.', 'warning')

            except Exception as e:
                flash(f'Eroare la citirea fisierului: {e}', 'danger')

    db.close()
    return render_template('stoc/import.html',
        locatie=locatie, matched=matched, unmatched=unmatched, importat=importat,
        today=datetime.now().strftime('%Y-%m-%d'))


def _parse_excel_stoc_matrice(file_storage):
    prefix_re = re.compile(r'^\d+\.\s*')
    continut = file_storage.read()
    wb = openpyxl.load_workbook(io.BytesIO(continut), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = rows[0]
    locatii_header = [str(v).strip() if v is not None else None for v in header[1:]]
    triplets = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        produs_str = prefix_re.sub('', str(row[0]).strip()).strip()
        if not produs_str:
            continue
        for i, val in enumerate(row[1:]):
            if i >= len(locatii_header) or not locatii_header[i]:
                continue
            if val is None:
                continue
            try:
                qty = float(str(val).replace(',', '.').strip())
            except (ValueError, TypeError):
                continue
            triplets.append((locatii_header[i], produs_str, qty))
    return locatii_header, triplets


def _fuzzy_match_loc(loc_str, loc_by_lower):
    loc_norm = re.sub(r'\s+', ' ', loc_str.strip().lower())
    if loc_norm in loc_by_lower:
        return loc_by_lower[loc_norm]
    for key, l in loc_by_lower.items():
        if loc_norm in key or key in loc_norm:
            return l
    words = [w for w in loc_norm.split() if len(w) > 4 and w not in ('carrefour', 'shopping')]
    if words:
        best_loc, best_score = None, 0
        for key, l in loc_by_lower.items():
            score = sum(1 for w in words if w in key)
            if score > best_score:
                best_score, best_loc = score, l
        if best_loc and best_score > 0:
            return best_loc
    matches = difflib.get_close_matches(loc_norm, loc_by_lower.keys(), n=1, cutoff=0.65)
    if matches:
        return loc_by_lower[matches[0]]
    return None


def _match_matrice(db, locatii_header, triplets):
    locatii = db.execute("SELECT id, nume FROM locatii WHERE activa=1").fetchall()
    produse = db.execute("SELECT id, denumire, cod_articol, cod_ean FROM produse WHERE activ=1").fetchall()
    loc_by_lower = {l['nume'].lower(): l for l in locatii}
    by_name = {p['denumire'].lower().strip(): p for p in produse}
    by_cod  = {p['cod_articol'].strip(): p for p in produse if p['cod_articol']}
    by_ean  = {str(p['cod_ean']).strip(): p for p in produse if p['cod_ean']}

    loc_cache  = {}
    loc_mapping = []
    for loc_str in locatii_header:
        if not loc_str:
            continue
        if loc_str not in loc_cache:
            loc_cache[loc_str] = _fuzzy_match_loc(loc_str, loc_by_lower)
        loc = loc_cache[loc_str]
        loc_mapping.append({'excel': loc_str, 'db': loc['nume'] if loc else None,
                            'db_id': loc['id'] if loc else None})

    matched = []
    unmatched_produse = set()
    for loc_str, produs_str, qty in triplets:
        loc = loc_cache.get(loc_str)
        if not loc:
            continue
        p = by_ean.get(produs_str) or by_cod.get(produs_str) or by_name.get(produs_str.lower())
        if not p:
            ident_low = produs_str.lower()
            for key, pp in by_name.items():
                if ident_low in key or key in ident_low:
                    p = pp
                    break
        if p:
            matched.append({'locatie_id': loc['id'], 'produs_id': p['id'], 'cantitate': qty})
        else:
            unmatched_produse.add(produs_str)

    return matched, list(sorted(unmatched_produse)), loc_mapping


@stoc_bp.route('/import-matrice', methods=['GET', 'POST'])
@login_required
def import_matrice():
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('stoc.index'))

    db = get_db()
    importat     = False
    nr_importate = 0
    unmatched_produse = []
    loc_mapping  = []

    if request.method == 'POST':
        fisier = request.files.get('fisier')
        data_stoc = request.form.get('data_stoc', '').strip() or datetime.now().strftime('%Y-%m-%d')
        if not fisier or not fisier.filename:
            flash('Selecteaza un fisier Excel.', 'danger')
        else:
            try:
                locatii_header, triplets = _parse_excel_stoc_matrice(fisier)
                matched, unmatched_produse, loc_mapping = _match_matrice(db, locatii_header, triplets)
                if matched:
                    for m in matched:
                        db.execute("""
                            INSERT OR REPLACE INTO stoc_initial_locatii
                                (locatie_id, produs_id, cantitate, data, utilizator_id)
                            VALUES (?,?,?,?,?)
                        """, (m['locatie_id'], m['produs_id'], m['cantitate'], data_stoc, current_user.id))
                    db.commit()
                    importat     = True
                    nr_importate = len(matched)
                    flash(f'{nr_importate} randuri importate cu succes pentru data {data_stoc}.', 'success')
                else:
                    flash('Niciun rand nu a putut fi importat. Verifica formatul fisierului.', 'warning')
            except Exception as e:
                flash(f'Eroare la citirea fisierului: {e}', 'danger')

    db.close()
    return render_template('stoc/import_matrice.html',
        importat=importat, nr_importate=nr_importate,
        unmatched_produse=unmatched_produse, loc_mapping=loc_mapping,
        today=datetime.now().strftime('%Y-%m-%d'))
