from flask import Blueprint, render_template, request, redirect, url_for, flash, Response
from flask_login import login_required, current_user
from database import get_db
from datetime import datetime
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

comenzi_bp = Blueprint('comenzi', __name__, url_prefix='/comenzi')


def _init_tables(db):
    db.execute("""
        CREATE TABLE IF NOT EXISTS comenzi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            locatie_id INTEGER NOT NULL,
            submisie_id INTEGER,
            data_comanda TEXT,
            status TEXT DEFAULT 'noua',
            obs TEXT,
            creat_de INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS comenzi_detalii (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            comanda_id INTEGER NOT NULL,
            produs_id INTEGER NOT NULL,
            cantitate REAL DEFAULT 0
        )
    """)
    db.commit()


STATUS_LABELS = {
    'noua':         ('Nouă',         'warning'),
    'in_procesare': ('În procesare', 'info'),
    'livrata':      ('Livrată',      'success'),
    'anulata':      ('Anulată',      'secondary'),
}

STATUS_FLOW = {
    'noua': 'in_procesare',
    'in_procesare': 'livrata',
}


def _build_query(filters):
    """Construieste query-ul de comenzi cu filtrele date."""
    q = """
        SELECT c.*,
               l.nume as locatie_nume,
               COUNT(cd.id) as nr_produse,
               SUM(cd.cantitate) as total_cant
        FROM comenzi c
        LEFT JOIN locatii l ON c.locatie_id = l.id
        LEFT JOIN comenzi_detalii cd ON cd.comanda_id = c.id
        WHERE 1=1
    """
    params = []
    if filters.get('status') and filters['status'] != 'toate':
        q += " AND c.status=?"
        params.append(filters['status'])
    if filters.get('locatie_id'):
        q += " AND c.locatie_id=?"
        params.append(int(filters['locatie_id']))
    if filters.get('data_de'):
        q += " AND c.data_comanda >= ?"
        params.append(filters['data_de'])
    if filters.get('data_pana'):
        q += " AND c.data_comanda <= ?"
        params.append(filters['data_pana'])
    q += " GROUP BY c.id ORDER BY c.data_comanda DESC, c.created_at DESC"
    return q, params


@comenzi_bp.route('/')
@login_required
def index():
    db = get_db()
    _init_tables(db)

    filters = {
        'status':    request.args.get('status', 'toate'),
        'locatie_id':request.args.get('locatie_id', ''),
        'data_de':   request.args.get('data_de', ''),
        'data_pana': request.args.get('data_pana', ''),
    }

    # Submisii tip comanda in asteptare (de validat)
    pending = db.execute("""
        SELECT s.*, l.nume as locatie_nume,
               (SELECT COUNT(*) FROM submisii_detalii sd WHERE sd.submisie_id=s.id) as nr_produse
        FROM submisii s
        LEFT JOIN locatii l ON s.locatie_id=l.id
        WHERE s.tip='comanda' AND s.status='pending'
        ORDER BY s.data_submisie ASC
    """).fetchall()

    q, params = _build_query(filters)
    comenzi = db.execute(q, params).fetchall()

    counts = {s: db.execute(
        "SELECT COUNT(*) FROM comenzi WHERE status=?", (s,)
    ).fetchone()[0] for s in STATUS_LABELS}
    counts['toate'] = sum(counts.values())

    locatii = db.execute(
        "SELECT id, nume FROM locatii WHERE activa=1 ORDER BY nume"
    ).fetchall()
    produse = db.execute(
        "SELECT id, denumire, unitate_masura FROM produse WHERE activ=1 ORDER BY denumire"
    ).fetchall()

    db.close()
    return render_template('comenzi/index.html',
                           comenzi=comenzi,
                           pending=pending,
                           filters=filters,
                           counts=counts,
                           status_labels=STATUS_LABELS,
                           locatii=locatii,
                           produse=produse,
                           today=datetime.now().strftime('%Y-%m-%d'))


@comenzi_bp.route('/valideaza/<int:sub_id>', methods=['POST'])
@login_required
def valideaza(sub_id):
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('comenzi.index'))

    db = get_db()
    _init_tables(db)

    sub = db.execute("SELECT * FROM submisii WHERE id=? AND tip='comanda' AND status='pending'",
                     (sub_id,)).fetchone()
    if not sub:
        flash('Submisie invalida sau deja procesata.', 'danger')
        db.close()
        return redirect(url_for('comenzi.index'))

    detalii = db.execute(
        "SELECT * FROM submisii_detalii WHERE submisie_id=? AND produs_id IS NOT NULL",
        (sub_id,)
    ).fetchall()

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data_azi = datetime.now().strftime('%Y-%m-%d')
    obs_admin = request.form.get('obs_admin', '').strip()

    cur = db.execute("""
        INSERT INTO comenzi (locatie_id, submisie_id, data_comanda, obs, creat_de, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (sub['locatie_id'], sub_id, data_azi, obs_admin, current_user.id, now))
    comanda_id = cur.lastrowid

    for d in detalii:
        db.execute("""
            INSERT INTO comenzi_detalii (comanda_id, produs_id, cantitate)
            VALUES (?, ?, ?)
        """, (comanda_id, d['produs_id'], d['cantitate']))

    db.execute("""
        UPDATE submisii SET status='aprobat', obs_admin=?, revizuit_de=?, revizuit_la=?
        WHERE id=?
    """, (obs_admin, current_user.id, now, sub_id))
    db.commit()
    db.close()

    flash('Comanda aprobată și salvată.', 'success')
    return redirect(url_for('comenzi.index'))


@comenzi_bp.route('/respinge/<int:sub_id>', methods=['POST'])
@login_required
def respinge(sub_id):
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('comenzi.index'))

    db = get_db()
    obs_admin = request.form.get('obs_admin', '').strip()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute("""
        UPDATE submisii SET status='respins', obs_admin=?, revizuit_de=?, revizuit_la=?
        WHERE id=? AND tip='comanda' AND status='pending'
    """, (obs_admin, current_user.id, now, sub_id))
    db.commit()
    db.close()
    flash('Comanda respinsă.', 'warning')
    return redirect(url_for('comenzi.index'))


@comenzi_bp.route('/preview-submisie/<int:sub_id>')
@login_required
def preview_submisie(sub_id):
    from flask import jsonify
    db = get_db()
    rows = db.execute("""
        SELECT p.denumire, p.unitate_masura as um, sd.cantitate
        FROM submisii_detalii sd
        JOIN produse p ON sd.produs_id = p.id
        WHERE sd.submisie_id = ? AND sd.produs_id IS NOT NULL
        ORDER BY p.denumire
    """, (sub_id,)).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@comenzi_bp.route('/export')
@login_required
def export():
    db = get_db()
    _init_tables(db)

    filters = {
        'status':    request.args.get('status', 'toate'),
        'locatie_id':request.args.get('locatie_id', ''),
        'data_de':   request.args.get('data_de', ''),
        'data_pana': request.args.get('data_pana', ''),
    }

    # Toate randurile detaliate (un rand per produs per comanda)
    q_detail = """
        SELECT c.id as comanda_id, l.nume as locatie_nume,
               c.data_comanda, c.status, c.obs,
               p.denumire as produs, p.unitate_masura as um,
               cd.cantitate
        FROM comenzi c
        LEFT JOIN locatii l ON c.locatie_id = l.id
        LEFT JOIN comenzi_detalii cd ON cd.comanda_id = c.id
        LEFT JOIN produse p ON cd.produs_id = p.id
        WHERE 1=1
    """
    params = []
    if filters['status'] and filters['status'] != 'toate':
        q_detail += " AND c.status=?"
        params.append(filters['status'])
    if filters['locatie_id']:
        q_detail += " AND c.locatie_id=?"
        params.append(int(filters['locatie_id']))
    if filters['data_de']:
        q_detail += " AND c.data_comanda >= ?"
        params.append(filters['data_de'])
    if filters['data_pana']:
        q_detail += " AND c.data_comanda <= ?"
        params.append(filters['data_pana'])
    q_detail += " ORDER BY c.data_comanda DESC, c.id, p.denumire"

    rows = db.execute(q_detail, params).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comenzi"

    # Stiluri
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1A3A5C")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_fills = {
        'noua':         PatternFill("solid", fgColor="FFF3CD"),
        'in_procesare': PatternFill("solid", fgColor="CFF4FC"),
        'livrata':      PatternFill("solid", fgColor="D1E7DD"),
        'anulata':      PatternFill("solid", fgColor="E2E3E5"),
    }
    status_ro = {
        'noua': 'Nouă', 'in_procesare': 'În procesare',
        'livrata': 'Livrată', 'anulata': 'Anulată',
    }

    headers = ['CMD #', 'Locație', 'Data comenzii', 'Status', 'Produs', 'UM', 'Cantitate', 'Observații']
    col_widths = [10, 28, 14, 14, 40, 6, 12, 25]

    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"

    for ri, r in enumerate(rows, 2):
        status_key = r['status'] or 'noua'
        sfill = status_fills.get(status_key)
        values = [
            f"CMD-{r['comanda_id']:04d}",
            r['locatie_nume'] or '',
            r['data_comanda'] or '',
            status_ro.get(status_key, status_key),
            r['produs'] or '',
            r['um'] or '',
            r['cantitate'] or 0,
            r['obs'] or '',
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if sfill:
                cell.fill = sfill
            if ci == 7:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = f"A1:H{len(rows)+1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"comenzi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'}
    )


@comenzi_bp.route('/adauga', methods=['POST'])
@login_required
def adauga():
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('comenzi.index'))

    db = get_db()
    _init_tables(db)

    locatie_id = request.form.get('locatie_id')
    data_comanda = request.form.get('data_comanda') or datetime.now().strftime('%Y-%m-%d')
    obs = request.form.get('obs', '').strip()

    if not locatie_id:
        flash('Selecteaza locatia.', 'danger')
        db.close()
        return redirect(url_for('comenzi.index'))

    cur = db.execute("""
        INSERT INTO comenzi (locatie_id, data_comanda, obs, creat_de, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (int(locatie_id), data_comanda, obs, current_user.id,
          datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    comanda_id = cur.lastrowid

    produse = db.execute("SELECT id FROM produse WHERE activ=1").fetchall()
    nr = 0
    for p in produse:
        val = request.form.get(f'cant_{p["id"]}', '').strip()
        try:
            cant = float(val.replace(',', '.')) if val else 0
        except ValueError:
            cant = 0
        if cant > 0:
            db.execute("""
                INSERT INTO comenzi_detalii (comanda_id, produs_id, cantitate)
                VALUES (?, ?, ?)
            """, (comanda_id, p['id'], cant))
            nr += 1

    if nr == 0:
        db.execute("DELETE FROM comenzi WHERE id=?", (comanda_id,))
        db.commit()
        db.close()
        flash('Introdu cel putin un produs in comanda.', 'warning')
        return redirect(url_for('comenzi.index'))

    db.commit()
    db.close()
    flash(f'Comanda adaugata cu {nr} produse.', 'success')
    return redirect(url_for('comenzi.detalii', comanda_id=comanda_id))


@comenzi_bp.route('/<int:comanda_id>')
@login_required
def detalii(comanda_id):
    db = get_db()
    _init_tables(db)

    c = db.execute("""
        SELECT c.*, l.nume as locatie_nume, u.nume_complet as creat_de_nume
        FROM comenzi c
        LEFT JOIN locatii l ON c.locatie_id = l.id
        LEFT JOIN utilizatori u ON c.creat_de = u.id
        WHERE c.id=?
    """, (comanda_id,)).fetchone()

    if not c:
        flash('Comanda negasita.', 'danger')
        db.close()
        return redirect(url_for('comenzi.index'))

    produse = db.execute("""
        SELECT cd.*, p.denumire, p.unitate_masura
        FROM comenzi_detalii cd
        JOIN produse p ON cd.produs_id = p.id
        WHERE cd.comanda_id = ?
        ORDER BY p.denumire
    """, (comanda_id,)).fetchall()

    db.close()
    return render_template('comenzi/detalii.html',
                           c=c,
                           produse=produse,
                           status_labels=STATUS_LABELS,
                           status_flow=STATUS_FLOW)


@comenzi_bp.route('/<int:comanda_id>/status', methods=['POST'])
@login_required
def update_status(comanda_id):
    if not current_user.is_manager():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('comenzi.index'))

    db = get_db()
    nou_status = request.form.get('status')
    if nou_status not in STATUS_LABELS:
        flash('Status invalid.', 'danger')
        db.close()
        return redirect(url_for('comenzi.detalii', comanda_id=comanda_id))

    db.execute("UPDATE comenzi SET status=? WHERE id=?", (nou_status, comanda_id))
    db.commit()
    db.close()
    flash(f'Status actualizat: {STATUS_LABELS[nou_status][0]}', 'success')
    return redirect(url_for('comenzi.detalii', comanda_id=comanda_id))


@comenzi_bp.route('/<int:comanda_id>/sterge', methods=['POST'])
@login_required
def sterge(comanda_id):
    if not current_user.is_admin():
        flash('Acces restrictionat.', 'danger')
        return redirect(url_for('comenzi.index'))
    db = get_db()
    db.execute("DELETE FROM comenzi_detalii WHERE comanda_id=?", (comanda_id,))
    db.execute("DELETE FROM comenzi WHERE id=?", (comanda_id,))
    db.commit()
    db.close()
    flash('Comanda stearsa.', 'success')
    return redirect(url_for('comenzi.index'))
