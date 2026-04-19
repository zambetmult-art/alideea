from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
from database import get_db, init_db
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font
import os
import socket

def get_lan_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'alideea-secret-2024')

with app.app_context():
    init_db()

ADMIN_USER = 'admin'
ADMIN_PASS = 'Sascanoua@12'

@app.context_processor
def inject_server_url():
    from flask import request as req
    host = req.host_url.rstrip('/')
    if 'localhost' in host or '127.0.0.1' in host:
        host = f'http://{get_lan_ip()}:5050'
    return {'server_url': host}

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def get_week_monday():
    today = date.today()
    monday = today - __import__('datetime').timedelta(days=today.weekday())
    return monday.strftime('%Y-%m-%d')

def count_unseen_orders():
    db = get_db()
    count = db.execute("SELECT COUNT(*) FROM orders WHERE seen=0").fetchone()[0]
    db.close()
    return count

# ─── AUTH ───────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['username'] == ADMIN_USER and request.form['password'] == ADMIN_PASS:
            session['admin'] = True
            return redirect(url_for('dashboard'))
        error = 'Date incorecte.'
    return render_template('admin/login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── ADMIN ──────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    db = get_db()
    unseen = count_unseen_orders()
    recent_orders = db.execute('''
        SELECT o.*, l.name as location_name
        FROM orders o JOIN locations l ON o.location_id = l.id
        ORDER BY o.submitted_at DESC LIMIT 10
    ''').fetchall()
    recent_inventory = db.execute('''
        SELECT ir.*, l.name as location_name
        FROM inventory_reports ir JOIN locations l ON ir.location_id = l.id
        ORDER BY ir.submitted_at DESC LIMIT 10
    ''').fetchall()
    total_products = db.execute("SELECT COUNT(*) FROM products WHERE active=1").fetchone()[0]
    total_locations = db.execute("SELECT COUNT(*) FROM locations WHERE active=1").fetchone()[0]
    db.close()
    return render_template('admin/dashboard.html',
        unseen=unseen,
        recent_orders=recent_orders,
        recent_inventory=recent_inventory,
        total_products=total_products,
        total_locations=total_locations
    )

@app.route('/api/unseen-orders')
@login_required
def api_unseen_orders():
    return jsonify({'count': count_unseen_orders()})

# ─── PRODUCTS ───────────────────────────────────────────────────────────────

@app.route('/produse')
@login_required
def products():
    db = get_db()
    products = db.execute("SELECT * FROM products ORDER BY name").fetchall()
    db.close()
    return render_template('admin/products.html', products=products, unseen=count_unseen_orders())

@app.route('/produse/adauga', methods=['POST'])
@login_required
def add_product():
    name = request.form['name'].strip()
    unit = request.form.get('unit', 'buc').strip()
    if name:
        db = get_db()
        db.execute("INSERT INTO products (name, unit) VALUES (?, ?)", (name, unit))
        db.commit()
        db.close()
        flash('Produs adăugat.', 'success')
    return redirect(url_for('products'))

@app.route('/produse/editeaza/<int:pid>', methods=['POST'])
@login_required
def edit_product(pid):
    name = request.form['name'].strip()
    unit = request.form.get('unit', 'buc').strip()
    active = 1 if request.form.get('active') else 0
    db = get_db()
    db.execute("UPDATE products SET name=?, unit=?, active=? WHERE id=?", (name, unit, active, pid))
    db.commit()
    db.close()
    flash('Produs actualizat.', 'success')
    return redirect(url_for('products'))

@app.route('/produse/sterge/<int:pid>', methods=['POST'])
@login_required
def delete_product(pid):
    db = get_db()
    db.execute("UPDATE products SET active=0 WHERE id=?", (pid,))
    db.commit()
    db.close()
    flash('Produs dezactivat.', 'info')
    return redirect(url_for('products'))

# ─── LOCATIONS ──────────────────────────────────────────────────────────────

@app.route('/locatii')
@login_required
def locations():
    db = get_db()
    locs = db.execute("SELECT * FROM locations ORDER BY name").fetchall()
    db.close()
    return render_template('admin/locations.html', locations=locs, unseen=count_unseen_orders())

@app.route('/locatii/adauga', methods=['POST'])
@login_required
def add_location():
    name = request.form['name'].strip()
    city = request.form.get('city', '').strip()
    if name:
        db = get_db()
        db.execute("INSERT INTO locations (name, city) VALUES (?, ?)", (name, city))
        db.commit()
        db.close()
        flash('Locație adăugată.', 'success')
    return redirect(url_for('locations'))

@app.route('/locatii/editeaza/<int:lid>', methods=['POST'])
@login_required
def edit_location(lid):
    name = request.form['name'].strip()
    city = request.form.get('city', '').strip()
    active = 1 if request.form.get('active') else 0
    db = get_db()
    db.execute("UPDATE locations SET name=?, city=?, active=? WHERE id=?", (name, city, active, lid))
    db.commit()
    db.close()
    flash('Locație actualizată.', 'success')
    return redirect(url_for('locations'))

# ─── ORDERS (ADMIN) ─────────────────────────────────────────────────────────

@app.route('/comenzi')
@login_required
def orders():
    db = get_db()
    db.execute("UPDATE orders SET seen=1")
    db.commit()
    orders = db.execute('''
        SELECT o.*, l.name as location_name
        FROM orders o JOIN locations l ON o.location_id = l.id
        ORDER BY o.submitted_at DESC
    ''').fetchall()
    db.close()
    db2 = get_db()
    locations = [dict(r) for r in db2.execute("SELECT id, name FROM locations WHERE active=1 ORDER BY name").fetchall()]
    db2.close()
    return render_template('admin/orders.html', orders=orders, unseen=0, locations_for_links=locations)

@app.route('/comenzi/<int:oid>')
@login_required
def order_detail(oid):
    db = get_db()
    order = db.execute('''
        SELECT o.*, l.name as location_name
        FROM orders o JOIN locations l ON o.location_id = l.id
        WHERE o.id=?
    ''', (oid,)).fetchone()
    items = db.execute('''
        SELECT oi.*, p.name as product_name, p.unit
        FROM order_items oi JOIN products p ON oi.product_id = p.id
        WHERE oi.order_id=? AND oi.quantity > 0
    ''', (oid,)).fetchall()
    db.close()
    return render_template('admin/order_detail.html', order=order, items=items, unseen=count_unseen_orders())

@app.route('/comenzi/<int:oid>/status', methods=['POST'])
@login_required
def update_order_status(oid):
    status = request.form['status']
    db = get_db()
    db.execute("UPDATE orders SET status=? WHERE id=?", (status, oid))
    db.commit()
    db.close()
    return redirect(url_for('order_detail', oid=oid))

# ─── INVENTORY (ADMIN) ──────────────────────────────────────────────────────

@app.route('/inventar')
@login_required
def inventory():
    db = get_db()
    reports = db.execute('''
        SELECT ir.*, l.name as location_name
        FROM inventory_reports ir JOIN locations l ON ir.location_id = l.id
        ORDER BY ir.submitted_at DESC
    ''').fetchall()
    db.close()
    db2 = get_db()
    locations = [dict(r) for r in db2.execute("SELECT id, name FROM locations WHERE active=1 ORDER BY name").fetchall()]
    db2.close()
    return render_template('admin/inventory.html', reports=reports, unseen=count_unseen_orders(), locations_for_links=locations)

@app.route('/inventar/<int:rid>')
@login_required
def inventory_detail(rid):
    db = get_db()
    report = db.execute('''
        SELECT ir.*, l.name as location_name
        FROM inventory_reports ir JOIN locations l ON ir.location_id = l.id
        WHERE ir.id=?
    ''', (rid,)).fetchone()
    items = db.execute('''
        SELECT ii.*, p.name as product_name, p.unit
        FROM inventory_items ii JOIN products p ON ii.product_id = p.id
        WHERE ii.report_id=?
        ORDER BY p.name
    ''', (rid,)).fetchall()
    db.close()
    return render_template('admin/inventory_detail.html', report=report, items=items, unseen=count_unseen_orders())

# ─── STOCK (ADMIN) ──────────────────────────────────────────────────────────

@app.route('/stoc')
@login_required
def stock():
    db = get_db()
    products = db.execute("SELECT * FROM products WHERE active=1 ORDER BY name").fetchall()
    stock_data = []
    for p in products:
        intrari = db.execute("SELECT COALESCE(SUM(quantity),0) FROM stock_entries WHERE product_id=?", (p['id'],)).fetchone()[0]
        iesiri = db.execute("SELECT COALESCE(SUM(quantity),0) FROM stock_exits WHERE product_id=?", (p['id'],)).fetchone()[0]
        stock_data.append({
            'id': p['id'],
            'name': p['name'],
            'unit': p['unit'],
            'intrari': intrari,
            'iesiri': iesiri,
            'stoc': intrari - iesiri
        })
    db.close()
    return render_template('admin/stock.html', stock_data=stock_data, unseen=count_unseen_orders(), today=date.today().strftime('%Y-%m-%d'))

@app.route('/stoc/intrare', methods=['POST'])
@login_required
def add_entry():
    product_id = request.form['product_id']
    quantity = float(request.form['quantity'])
    entry_date = request.form['date']
    notes = request.form.get('notes', '')
    db = get_db()
    db.execute("INSERT INTO stock_entries (product_id, quantity, date, notes) VALUES (?,?,?,?)",
               (product_id, quantity, entry_date, notes))
    db.commit()
    db.close()
    flash('Intrare adăugată.', 'success')
    return redirect(url_for('stock'))

@app.route('/stoc/iesire', methods=['POST'])
@login_required
def add_exit():
    product_id = request.form['product_id']
    quantity = float(request.form['quantity'])
    exit_date = request.form['date']
    notes = request.form.get('notes', '')
    db = get_db()
    db.execute("INSERT INTO stock_exits (product_id, quantity, date, notes) VALUES (?,?,?,?)",
               (product_id, quantity, exit_date, notes))
    db.commit()
    db.close()
    flash('Ieșire adăugată.', 'success')
    return redirect(url_for('stock'))

# ─── STORE LINK ─────────────────────────────────────────────────────────────

@app.route('/magazin')
def store_index():
    db = get_db()
    locations = db.execute("SELECT * FROM locations WHERE active=1 ORDER BY name").fetchall()
    db.close()
    return render_template('store/index.html', locations=locations)

@app.route('/magazin/inventar', methods=['GET', 'POST'])
def store_inventory():
    db = get_db()
    week_date = get_week_monday()

    if request.method == 'POST':
        location_id = request.form['location_id']
        # Blocare re-trimitere in aceeasi saptamana
        existing = db.execute(
            "SELECT id FROM inventory_reports WHERE location_id=? AND week_date=?",
            (location_id, week_date)
        ).fetchone()
        if existing:
            report = db.execute("SELECT ir.*, l.name as location_name FROM inventory_reports ir JOIN locations l ON ir.location_id=l.id WHERE ir.id=?", (existing['id'],)).fetchone()
            items = db.execute("SELECT ii.*, p.name as product_name, p.unit FROM inventory_items ii JOIN products p ON ii.product_id=p.id WHERE ii.report_id=? ORDER BY p.name", (existing['id'],)).fetchall()
            db.close()
            return render_template('store/inventory_view.html', report=report, items=items, already_sent=True)

        db.execute("INSERT INTO inventory_reports (location_id, week_date) VALUES (?,?)", (location_id, week_date))
        report_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        products = db.execute("SELECT * FROM products WHERE active=1 ORDER BY name").fetchall()
        for p in products:
            qty_str = request.form.get(f'qty_{p["id"]}', '0').strip()
            qty = float(qty_str) if qty_str else 0
            db.execute("INSERT INTO inventory_items (report_id, product_id, quantity) VALUES (?,?,?)",
                       (report_id, p['id'], qty))
        db.commit()
        report = db.execute("SELECT ir.*, l.name as location_name FROM inventory_reports ir JOIN locations l ON ir.location_id=l.id WHERE ir.id=?", (report_id,)).fetchone()
        items = db.execute("SELECT ii.*, p.name as product_name, p.unit FROM inventory_items ii JOIN products p ON ii.product_id=p.id WHERE ii.report_id=? ORDER BY p.name", (report_id,)).fetchall()
        db.close()
        return render_template('store/inventory_view.html', report=report, items=items, already_sent=False)

    location_id = request.args.get('location_id')
    # Daca au trimis deja saptamana asta, arata view-ul
    if location_id:
        existing = db.execute(
            "SELECT id FROM inventory_reports WHERE location_id=? AND week_date=?",
            (location_id, week_date)
        ).fetchone()
        if existing:
            report = db.execute("SELECT ir.*, l.name as location_name FROM inventory_reports ir JOIN locations l ON ir.location_id=l.id WHERE ir.id=?", (existing['id'],)).fetchone()
            items = db.execute("SELECT ii.*, p.name as product_name, p.unit FROM inventory_items ii JOIN products p ON ii.product_id=p.id WHERE ii.report_id=? ORDER BY p.name", (existing['id'],)).fetchall()
            db.close()
            return render_template('store/inventory_view.html', report=report, items=items, already_sent=True)

    locations = db.execute("SELECT * FROM locations WHERE active=1 ORDER BY name").fetchall()
    products = db.execute("SELECT * FROM products WHERE active=1 ORDER BY name").fetchall()
    db.close()
    return render_template('store/inventory.html', locations=locations, products=products, location_id=location_id)

@app.route('/magazin/comanda', methods=['GET', 'POST'])
def store_order():
    db = get_db()
    today = date.today().strftime('%Y-%m-%d')

    if request.method == 'POST':
        location_id = request.form['location_id']
        notes = request.form.get('notes', '')
        # Blocare re-trimitere in aceeasi zi
        existing = db.execute(
            "SELECT id FROM orders WHERE location_id=? AND date(submitted_at)=?",
            (location_id, today)
        ).fetchone()
        if existing:
            order = db.execute("SELECT o.*, l.name as location_name FROM orders o JOIN locations l ON o.location_id=l.id WHERE o.id=?", (existing['id'],)).fetchone()
            items = db.execute("SELECT oi.*, p.name as product_name, p.unit FROM order_items oi JOIN products p ON oi.product_id=p.id WHERE oi.order_id=? AND oi.quantity>0", (existing['id'],)).fetchall()
            db.close()
            return render_template('store/order_view.html', order=order, items=items, already_sent=True)

        products = db.execute("SELECT * FROM products WHERE active=1 ORDER BY name").fetchall()
        has_items = any(float(request.form.get(f'qty_{p["id"]}', '0') or '0') > 0 for p in products)
        if not has_items:
            locations = db.execute("SELECT * FROM locations WHERE active=1 ORDER BY name").fetchall()
            db.close()
            return render_template('store/order.html', locations=locations, products=products,
                                   location_id=location_id, error='Introduceți cel puțin o cantitate.')
        db.execute("INSERT INTO orders (location_id, notes) VALUES (?,?)", (location_id, notes))
        order_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        for p in products:
            qty_str = request.form.get(f'qty_{p["id"]}', '0').strip()
            qty = float(qty_str) if qty_str else 0
            if qty > 0:
                db.execute("INSERT INTO order_items (order_id, product_id, quantity) VALUES (?,?,?)",
                           (order_id, p['id'], qty))
        db.commit()
        order = db.execute("SELECT o.*, l.name as location_name FROM orders o JOIN locations l ON o.location_id=l.id WHERE o.id=?", (order_id,)).fetchone()
        items = db.execute("SELECT oi.*, p.name as product_name, p.unit FROM order_items oi JOIN products p ON oi.product_id=p.id WHERE oi.order_id=? AND oi.quantity>0", (order_id,)).fetchall()
        db.close()
        return render_template('store/order_view.html', order=order, items=items, already_sent=False)

    location_id = request.args.get('location_id')
    # Daca au trimis deja azi, arata view-ul
    if location_id:
        existing = db.execute(
            "SELECT id FROM orders WHERE location_id=? AND date(submitted_at)=?",
            (location_id, today)
        ).fetchone()
        if existing:
            order = db.execute("SELECT o.*, l.name as location_name FROM orders o JOIN locations l ON o.location_id=l.id WHERE o.id=?", (existing['id'],)).fetchone()
            items = db.execute("SELECT oi.*, p.name as product_name, p.unit FROM order_items oi JOIN products p ON oi.product_id=p.id WHERE oi.order_id=? AND oi.quantity>0", (existing['id'],)).fetchall()
            db.close()
            return render_template('store/order_view.html', order=order, items=items, already_sent=True)

    locations = db.execute("SELECT * FROM locations WHERE active=1 ORDER BY name").fetchall()
    products = db.execute("SELECT * FROM products WHERE active=1 ORDER BY name").fetchall()
    db.close()
    return render_template('store/order.html', locations=locations, products=products, location_id=location_id)

# ─── EXCEL EXPORT ───────────────────────────────────────────────────────────

def _excel_response(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _bold_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)

@app.route('/export/produse')
@login_required
def export_products():
    db = get_db()
    rows = db.execute("SELECT * FROM products ORDER BY name").fetchall()
    db.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produse"
    ws.append(["#", "Nume produs", "Unitate masura", "Status"])
    _bold_header(ws)
    for i, p in enumerate(rows, 1):
        ws.append([i, p['name'], p['unit'], 'Activ' if p['active'] else 'Inactiv'])
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    return _excel_response(wb, 'produse.xlsx')

@app.route('/export/stoc')
@login_required
def export_stock():
    db = get_db()
    products = db.execute("SELECT * FROM products WHERE active=1 ORDER BY name").fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stoc"
    ws.append(["Produs", "UM", "Intrari", "Iesiri", "Stoc"])
    _bold_header(ws)
    for p in products:
        intrari = db.execute("SELECT COALESCE(SUM(quantity),0) FROM stock_entries WHERE product_id=?", (p['id'],)).fetchone()[0]
        iesiri = db.execute("SELECT COALESCE(SUM(quantity),0) FROM stock_exits WHERE product_id=?", (p['id'],)).fetchone()[0]
        ws.append([p['name'], p['unit'], intrari, iesiri, intrari - iesiri])
    db.close()
    ws.column_dimensions['A'].width = 40
    return _excel_response(wb, 'stoc.xlsx')

@app.route('/export/comenzi')
@login_required
def export_orders():
    db = get_db()
    orders = db.execute('''
        SELECT o.*, l.name as location_name FROM orders o
        JOIN locations l ON o.location_id = l.id ORDER BY o.submitted_at DESC
    ''').fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comenzi"
    ws.append(["ID", "Locatie", "Data", "Status", "Produs", "Cantitate", "UM"])
    _bold_header(ws)
    for o in orders:
        items = db.execute('''
            SELECT oi.*, p.name as product_name, p.unit FROM order_items oi
            JOIN products p ON oi.product_id = p.id
            WHERE oi.order_id=? AND oi.quantity > 0
        ''', (o['id'],)).fetchall()
        if items:
            for item in items:
                ws.append([o['id'], o['location_name'], o['submitted_at'][:16], o['status'],
                           item['product_name'], item['quantity'], item['unit']])
        else:
            ws.append([o['id'], o['location_name'], o['submitted_at'][:16], o['status'], '', '', ''])
    db.close()
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['E'].width = 35
    return _excel_response(wb, 'comenzi.xlsx')

@app.route('/export/inventar')
@login_required
def export_inventory_excel():
    db = get_db()
    reports = db.execute('''
        SELECT ir.*, l.name as location_name FROM inventory_reports ir
        JOIN locations l ON ir.location_id = l.id ORDER BY ir.submitted_at DESC
    ''').fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventare"
    ws.append(["ID", "Locatie", "Saptamana", "Data trimiterii", "Produs", "Cantitate", "UM"])
    _bold_header(ws)
    for r in reports:
        items = db.execute('''
            SELECT ii.*, p.name as product_name, p.unit FROM inventory_items ii
            JOIN products p ON ii.product_id = p.id WHERE ii.report_id=? ORDER BY p.name
        ''', (r['id'],)).fetchall()
        if items:
            for item in items:
                ws.append([r['id'], r['location_name'], r['week_date'], r['submitted_at'][:16],
                           item['product_name'], item['quantity'], item['unit']])
        else:
            ws.append([r['id'], r['location_name'], r['week_date'], r['submitted_at'][:16], '', '', ''])
    db.close()
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 35
    return _excel_response(wb, 'inventar.xlsx')

# ─── EXCEL IMPORT ───────────────────────────────────────────────────────────

@app.route('/import/produse', methods=['POST'])
@login_required
def import_products():
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Niciun fisier selectat.', 'danger')
        return redirect(url_for('products'))
    try:
        wb = openpyxl.load_workbook(f)
    except Exception:
        flash('Fisier Excel invalid.', 'danger')
        return redirect(url_for('products'))
    ws = wb.active
    db = get_db()
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ''
        unit = str(row[1]).strip() if len(row) > 1 and row[1] else 'buc'
        if name and name.lower() != 'none':
            existing = db.execute("SELECT id FROM products WHERE name=?", (name,)).fetchone()
            if not existing:
                db.execute("INSERT INTO products (name, unit) VALUES (?, ?)", (name, unit))
                added += 1
    db.commit()
    db.close()
    flash(f'{added} produse importate.', 'success')
    return redirect(url_for('products'))

@app.route('/import/stoc', methods=['POST'])
@login_required
def import_stock():
    f = request.files.get('file')
    tip = request.form.get('tip', 'intrare')
    if not f or not f.filename:
        flash('Niciun fisier selectat.', 'danger')
        return redirect(url_for('stock'))
    try:
        wb = openpyxl.load_workbook(f)
    except Exception:
        flash('Fisier Excel invalid.', 'danger')
        return redirect(url_for('stock'))
    ws = wb.active
    db = get_db()
    added = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:
            continue
        product_name = str(row[0]).strip()
        try:
            quantity = float(row[1]) if row[1] else 0
        except Exception:
            errors.append(f'Randul {i}: cantitate invalida')
            continue
        if hasattr(row[2], 'strftime'):
            entry_date = row[2].strftime('%Y-%m-%d')
        else:
            entry_date = str(row[2]).strip() if len(row) > 2 and row[2] else date.today().strftime('%Y-%m-%d')
        notes = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        product = db.execute("SELECT id FROM products WHERE name=?", (product_name,)).fetchone()
        if not product:
            errors.append(f'Randul {i}: produsul "{product_name}" nu exista')
            continue
        if quantity > 0:
            if tip == 'intrare':
                db.execute("INSERT INTO stock_entries (product_id, quantity, date, notes) VALUES (?,?,?,?)",
                           (product['id'], quantity, entry_date, notes))
            else:
                db.execute("INSERT INTO stock_exits (product_id, quantity, date, notes) VALUES (?,?,?,?)",
                           (product['id'], quantity, entry_date, notes))
            added += 1
    db.commit()
    db.close()
    msg = f'{added} inregistrari importate.'
    if errors:
        msg += ' Erori: ' + '; '.join(errors[:5])
    flash(msg, 'success' if not errors else 'warning')
    return redirect(url_for('stock'))

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5050))
    app.run(debug=False, host='0.0.0.0', port=port)
